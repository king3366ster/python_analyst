# -*- coding:utf-8 -*- 
import pdb, re, datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy import *
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.dialects.mysql import DOUBLE, TIMESTAMP


def checkparams (func):
    def _checkparams (cmdobj, config = None, cache = None):
        ctype = cmdobj['ctype']
        if cache is None or (not isinstance(cache, dict)):
            raise Exception('Runtime Error: %s without cache' % ctype)
        if 'src' not in cmdobj['ckeys']:
            raise Exception('Command Error: %s without src' % ctype)
        src = cmdobj['ckeys']['src']
        if src not in cache:
            raise Exception('Runtime Error: %s %s not in cache' % (ctype, src))
        if 'tar' not in cmdobj['ckeys']:
            cmdobj['ckeys']['tar'] = cmdobj['ckeys']['src']
        return func(cmdobj, config = config, cache = cache)
    return _checkparams

def checkdb (func):
    def _checkdb (cmdobj, config = None, cache = None):
        ctype = cmdobj['ctype']
        if config is None:
            raise Exception('Config Error: %s without configs' % ctype)
        if 'db' not in cmdobj['ckeys']:
            raise Exception('Command Error: %s without db' % ctype)
        db = cmdobj['ckeys']['db']
        if db not in config:
            raise Exception('Command Error: %s without correct db' % ctype)
        return func(cmdobj, config = config, cache = cache)
    return _checkdb

@checkparams
def saveexcel (cmdobj, config = None, cache = None):
    ckeys = cmdobj['ckeys']
    src = ckeys['src']
    tar = ckeys['tar']
    if not re.search(r'\.xlsx$', tar):
        filename = tar + '.xlsx'
    else:
        filename = tar

    if 'if_exists' in ckeys:
        if_exists = ckeys['if_exists']
    else:
        if_exists = 'replace'
    if 'sheet' in ckeys:
        sheet_name = ckeys['sheet']
    else:
        sheet_name = 'Sheet1'
    data = cache[src]

    if if_exists == 'append':
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.get_sheet_by_name(sheet_name)
        except Exception as what:
            wb = openpyxl.workbook.Workbook()
            ws = wb.create_sheet(sheet_name, 0)
    else:
        wb = openpyxl.workbook.Workbook()
        ws = wb.create_sheet(sheet_name, 0)

    dfIndexs = data.index
    dfColumns = data.columns
    dfValues = data.values

    for i in range(0, len(dfColumns)):
        index_col = i + 1
        index_row = 1
        col = get_column_letter(index_col + 1)
        ws['%s%s' %(col, index_row)] = dfColumns[i]

    for i in range(0, len(dfIndexs)):
        index_row = i + 2 #dfIndexs[i] + 2
        ws['A%d' % index_row] = int(dfIndexs[i])
        for j in range(0, len(dfColumns)):
            index_col = j + 1
            col = get_column_letter(index_col + 1)
            tmpValue = dfValues[i][j]
            if pd.isnull(tmpValue):
                continue
            try:
                ws['%s%s' %(col, index_row)] = tmpValue
            except Exception as what:
                print what
    wb.save(filename = filename)

@checkparams
def savecsv (cmdobj, config = None, cache = None):
    ckeys = cmdobj['ckeys']
    src = ckeys['src']
    tar = ckeys['tar']
    if not re.search(r'\.csv', tar):
        filename = tar + '.csv'
    else:
        filename = tar
    cache[src].to_csv(filename, index = False, encoding = 'gbk')
    return cache[src]

def gen_engine_mysql (config):
    if 'host' in config:
        _host = config['host']
    else:
        _host = '127.0.0.1'
    if 'user' in config:
        _user = config['user']
    else:
        _user = 'root'
    if 'pwd' in config:
        _pwd = config['pwd']
    else:
        _pwd = '123456'
    if 'db' in config:
        _db = config['db']
    else:
        _db = 'mysql'
    if 'port' in config:
        _port = config['port']
    else:
        _port = 3306
    mysql_engine = 'mysql://%s:%s@%s:%s/%s?charset=utf8' % (_user, _pwd, _host, _port, _db)
    # print (mysql_engine)
    return create_engine(mysql_engine)

def create_mysql_table (mysql_engine, data, tb_name = None, unique_key_set = set(), need_datetime = 1):
    metadata = MetaData()
    created_at = False
    updated_at = False

    columns = [tb_name, metadata, Column('id', BIGINT, primary_key = True)]
    for col_name in data.columns:
        if col_name == 'id':
            continue
        col_type = unicode(data[col_name].dtype)
        temp_type = String(128)
        if col_type.find('int') >= 0:
            temp_type = Integer
        elif col_type.find('float') >= 0:
            temp_type = DOUBLE
        elif col_type.find('datetime') >= 0:
            temp_type = DateTime
        else:
            if re.search(r'id$', col_name):
                temp_type = BIGINT
            elif re.search(r'(updated|created)', col_name):
                temp_type = TIMESTAMP
            else:
                temp_type = String(128) 

        if col_name in unique_key_set and len(unique_key_set) == 1:
            columns.append(Column(col_name, temp_type, unique = True, nullable = False, autoincrement = False)) 
        elif col_name in unique_key_set:
            columns.append(Column(col_name, temp_type, nullable = False)) 
        else:
            columns.append(Column(col_name, temp_type))

        if col_name == 'created_at':
            created_at = True
        if col_name == 'updated_at':
            updated_at = True

    if need_datetime == True:
        if updated_at == False:
            columns.append(Column('updated_at', TIMESTAMP, nullable = False, server_default = text('CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP'))) 
        if created_at == False:
            columns.append(Column('created_at', TIMESTAMP, nullable = False))
    if len(unique_key_set) > 1:
        columns.append(apply(UniqueConstraint, list(unique_key_set)))
    tb_create = apply(Table, columns)   
    metadata.create_all(mysql_engine)

@checkparams
@checkdb
def savemysql (cmdobj, config = None, cache = None):
    cmdkeys = cmdobj['ckeys']
    src = cmdkeys['src']
    data = cache[src]

    tb_name = cmdkeys['tar']
    db = cmdkeys['db']
    mysql_engine = gen_engine_mysql(config[db])
    DB_Session = sessionmaker(bind = mysql_engine)
    mysql_session = DB_Session()
    
    if 'if_exists' in cmdkeys:
        if_exists = cmdkeys['if_exists']
    else:
        if_exists = 'append'

    if 'need_datetime' in cmdkeys:
        need_datetime = cmdkeys['need_datetime']
    else:
        need_datetime = 1

    if 'unique' in cmdkeys:
        unique_key = cmdkeys['unique']
        unique_keys = re.split(r'\s+', unique_key)
    else:
        unique_key = None
        unique_keys = []
        if if_exists == 'append' or if_exists == 'append_ignore' or if_exists == 'replace':
            if if_exists == 'append_ignore':
                if_exists = 'append'
            data.to_sql(tb_name, mysql_engine, if_exists = if_exists, index = False)
            return data
        else:
            raise Exception('Command Error: savemysql if_exists illegal')

    unique_key_set = set()
    for ukey in unique_keys:
        unique_key_set.add(ukey)

    # create table if not exists
    tb_count = mysql_session.execute('SHOW TABLES LIKE "%%%s%%"' % tb_name).first()
    if tb_count is None:
        create_mysql_table(mysql_engine, data, tb_name = tb_name, unique_key_set = unique_key_set, need_datetime = need_datetime)

    # 获取表结构
    metadata = MetaData(mysql_engine)
    tb_model = Table(tb_name, metadata, autoload = True)

    # 此处去除数据库中表不存在的列
    tb_column_set = set()
    for col_name in tb_model.c:
        tb_column_set.add(unicode(col_name).replace('%s.' % tb_name, ''))
    dataset = data.copy(deep = True)
    for col_name in dataset.columns:
        if unicode(col_name) not in tb_column_set:
            dataset.drop(col_name, axis = 1, inplace = True)
    # uniquekey与columns交集
    unique_key_set = unique_key_set & tb_column_set
    if len(unique_key_set) <= 0:
        raise Exception('Command Error: savemysql all unique keys not in target table')
    # 清空数据库，如果选项为replace
    if if_exists == 'replace':
        tb_model.delete().execute()
    key_col = list(unique_key_set)[0]

    dtime = datetime.datetime.now()
    for value in dataset.iterrows():
        value_illegal = False
        insert_value = value[1].to_dict()
        update_value = dict()
        for vkey in insert_value:
            value = insert_value[vkey]
            if not pd.isnull(value):
                update_value[vkey] = value
        for ukey in unique_key_set:
            if ukey not in update_value:
                value_illegal = True
        if value_illegal:
            print ('Runtime Error: mysql insert with unique key null, ignore')
            continue
        
        if 'updated_at' not in update_value and 'updated_at' in tb_column_set:
            update_value['updated_at'] = dtime

        if if_exists == 'replace':
            if 'created_at' not in update_value and 'created_at' in tb_column_set:
                update_value['created_at'] = dtime
            try:
                tb_model.insert(values = update_value).execute()
            except Exception as what: # 一般为 duplicate entry
                if unicode(what).find('Duplicate entry') >= 0:
                    pass
                else:
                    print (unicode(what))
        else:
            ukey_compare = []
            for ukey in unique_key_set:
                tmpval = update_value[ukey]
                if isinstance(tmpval, unicode) or isinstance(tmpval, str):
                    tmpval = tmpval.replace('\"', '\\\"')
                    tmpval = '\"%s\"' % tmpval
                else:
                    tmpval = unicode(tmpval)
                ukey_compare.append('`%s`=%s' % (ukey, tmpval))
            where_cause = ' AND '.join(ukey_compare)
            tb_count = mysql_session.execute('SELECT COUNT(1) FROM `%s` WHERE %s' % (tb_name, where_cause)).first()
            if tb_count[0] > 0: # 记录已存在
                if if_exists == 'append_ignore':
                    continue
                else:
                    try:
                        tb_model.update(values = update_value).where(where_cause).execute()
                    except Exception as what:
                        print (unicode(what))
            else:
                if 'created_at' not in update_value and 'created_at' in tb_column_set:
                    update_value['created_at'] = dtime
                try:
                    tb_model.insert(values = update_value).execute()
                except Exception as what:
                    print (unicode(what))
    mysql_session.close()




    